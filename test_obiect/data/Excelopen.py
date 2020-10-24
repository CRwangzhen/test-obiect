from openpyxl import *
from config.moudler_log import  logger1

class excelData:
    def getExcel(self):
        a = logger1()
        #取workbook
        workbook=load_workbook(r"C:\Users\10466\Desktop\新建文件夹\pytest_file\test_obiect\data\case.xlsx")
        workbook.sheetnames
        # wb.get_sheet_names()

        # print(workbook)
        #取sheet
        sheet=workbook["Sheet1"]
        #定义外层的list结构
        lists=[]
        #读取rows
        row_sheet=sheet.iter_rows()
        #循环读取每一行，需要赋值每一行为一个list
        for item in row_sheet:
            #如果取到第一行就跳出去直接取下一行
            if item[0].value == "url":
                continue
            list=[]
            for col in item:

                list.append(col.value)
            # print(list)
            list = tuple(list)
            # print("xxxxx",list)
            lists.append(list)


        a.applog_pen(lists)
        print(lists)
        return lists
# A = excelData().getExcel()
# print("AAAAAAAAAAAAAAA",A)
if __name__ == "__main__":
    excelData().getExcel()




